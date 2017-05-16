import openpyxl
import ffile

def make_excel_file(total_data_dic, filename, folder):
    """
    Gets a dic of lists of lists of data to output into cells.
    Each lists will give a new row, and each element in the lists
        will create a new cell in the next column.
    The key of the dic will be the sheetname, while the prop will
        be first level of lists.
    :param data_list: dictionary of lists of lists of data
    :param filename: string of the excel filename
    :param folder: string of where the folder will be
    :return: None
    """
    ffile.move_dir(folder)

    wb = openpyxl.Workbook()

    for sheetname, excel_data_lists in total_data_dic.items():
        sheet = wb.create_sheet(title = sheetname)
        for y, row_data_lists in enumerate(excel_data_lists):
            rowNum = y + 1
            for x, data in enumerate(row_data_lists):
                columnNum = x + 1
                sheet.cell(row=rowNum, column=columnNum).value = data

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    wb.save(filename)

    ffile.dir_back()