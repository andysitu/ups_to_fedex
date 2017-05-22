from . import excel_helper
import ffile
import openpyxl
import re, datetime

max_earned_discount_num = 5

def process_excel_fedex(earned_num = 0):
    """
    Input: earned_num(float), signifying the earned discount obtained.
    Output: rate_dic  -> [ sheet/ delivery name ] [ weight(dic) ][zone] = rate(float) 
    """
    ffile.move_dir("fedex_rates")

    earned_string = ''

    if earned_num == 0:
        earned_string = 'no'
    elif earned_num == 1:
        earned_string = '1st'
    elif earned_num == 2:
        earned_string = '2nd'
    elif earned_num == 3:
        earned_string = '3rd'
    elif earned_num == 4:
        earned_string = '4th'
    elif earned_num == 5:
        earned_string = '5th'

    rate_dic = {}

    fedex_rate_excel_filename = "fedex_discounted_rates_" + earned_string + "_earned.xlsx"

    wb = openpyxl.load_workbook(fedex_rate_excel_filename)
    sheet_names_list = wb.get_sheet_names()

    for sheet_name in sheet_names_list:
        sheet = wb.get_sheet_by_name(sheet_name)
        rates = proc_sheet_for_rates(sheet)
        rate_dic[sheet_name] = rates

    ffile.dir_back()

    return rate_dic

def get_all_fedex_rates():
    total_rate_dic = {}
    for earned_discount_num in range(max_earned_discount_num + 1):
        fedex_rate_dic = process_excel_fedex(earned_discount_num)
        total_rate_dic[earned_discount_num] = fedex_rate_dic
    return total_rate_dic

def proc_sheet_for_rates(sheet):
    """
    Input: sheet from openpyxl
    Output: Dictionary -> weight_dic [ weight(dic) ][zone] = rate(float)
    """
    zone_row_num = excel_helper.zone_row
    max_rows = sheet.max_row
    max_columns = sheet.max_column

    col_letters_list = []
    for i in range(2, max_columns + 1):
        col_letters_list.append( excel_helper.rev_alphabet_dic[i] )

    zones_dic = excel_helper.get_zones(sheet)

    ##Start translate rows and weight for rate
    weight_dic = {}
    for row in range(excel_helper.rate_start_row, max_rows + 1):
        weight_cell_loc = excel_helper.weight_column + str(row)
        weight_str = sheet[weight_cell_loc].value
        weight = ''
        try:
            weight = int(weight_str)
        except ValueError:
            weight = weight_str
        weight_dic[weight] = {}

        for letter, zone_dic in zones_dic.items():
            if letter == excel_helper.weight_column:
                continue
            cell_loc = letter + str(row)
            zone = int(zone_dic["start"])
            rate = sheet[cell_loc].value
            weight_dic[weight][zone] = rate
    return weight_dic

def get_fedex_calc_function(fedex_charge_type):
    if fedex_charge_type == "Ground":
        return calc_ground_commercial
    elif fedex_charge_type == "Fuel Surcharge":
        return calc_fuel_surcharge
    elif fedex_charge_type == "Additional Handling Surcharge":
        return calc_add_handling
    elif fedex_charge_type == "Home Delivery":
        return calc_ground_residential
    elif fedex_charge_type == "Residential Delivery Charge":
        return calc_residential_charge
    elif fedex_charge_type == "Delivery Area Surcharge":
        return calc_delivery_area_surcharge
    elif fedex_charge_type == "Delivery Signature":
        return calc_signature
    elif fedex_charge_type == 'Smart Post 1-70 lbs':
        return calc_smart_post_1lb_plus
    elif fedex_charge_type == "Oversize Charge":
        return calc_oversize_charge
    elif fedex_charge_type == "Non-Machinable":
        return calc_nonmachinable_charge
    elif fedex_charge_type == "2 Day":
        return calc_2_day

# Calculate the rates for fedex charges
def get_rate(fedex_service_name, weight, zone, rates_dic):
    weight = int(weight)
    zone = int(zone)
    try:
        return rates_dic[fedex_service_name][weight][zone]
    except KeyError as e:
        print(rates_dic[fedex_service_name])
        print("KEY ERROR in get_rate")
        print(fedex_service_name)
        print(weight)
        print(zone)
        raise(e)


def calc_ground_commercial(weight, zone, rates_dic):
    return get_rate('Ground', weight, zone, rates_dic)


def calc_ground_residential(weight, zone, rates_dic):
    return calc_ground_commercial(weight, zone, rates_dic)


def calc_2_day(weight, zone, rates_dic):
    return get_rate('2 Day', weight, zone, rates_dic)


def calc_smart_post_1lb_plus(weight, zone, rates_dic):
    return get_rate('Smart Post 1-70 lbs', weight, zone, rates_dic)


def calc_residential_charge(weight, zone):
    res_surcharge = 3.45
    discount = 0.50
    return res_surcharge - discount


def calc_oversize_charge(weight, zone):
    oversize_charge = 72.50
    discount = 0.00
    return oversize_charge - discount


def calc_add_handling(weight, zone):
    add_handling = 11.00
    discount = add_handling * 0.25
    return add_handling - discount


def calc_delivery_area_surcharge(service_type, residential, extended):
    # type refers to residential or commercial
    delivery_area_surcharge = 0
    if service_type == "Ground":
        if residential:
            if extended:
                delivery_area_surcharge = 4.2
            else:
                delivery_area_surcharge = 3.9
        else:
            delivery_area_surcharge = 2.45
    elif service_type == 'Priority Overnight' or service_type == 'Standard Overnight' or service_type == '2 Day AM' or service_type == '2 Day':
        if residential:
            if extended:
                delivery_area_surcharge = 4.2
            else:
                delivery_area_surcharge = 3.9
        else:
            delivery_area_surcharge = 2.6
    elif service_type == "Home Delivery":
        if extended:
            delivery_area_surcharge = 4.2
        else:
            delivery_area_surcharge = 3.35
    elif service_type == 'Smart Post 1-16 oz' or service_type == 'Smart Post 1-70 lbs':
        if extended:
            delivery_area_surcharge = 1.50
        else:
            delivery_area_surcharge = 1.00
    else:
        msg = "Unknown service_type " + service_type
        print(msg)

    discount = delivery_area_surcharge * 0.25

    return delivery_area_surcharge - discount


def calc_signature(weight, zone):
    signature_rate = 4.5
    discount = signature_rate * 0.25
    return signature_rate - discount


def calc_nonmachinable_charge(weight, zone):
    nonmachinable_charge = 2.5
    return nonmachinable_charge

add_fuel_surcharge_index = {
	# Charge types with True are calculated in fuel surcharge percentage
		"Ground": True,
		"Home Delivery": True,
		"Residential Delivery Charge": True,
		"Delivery Area Surcharge": True,
		'Smart Post 1-70 lbs': True,
		"Oversize Charge": True,
		"Non-Machinable": False,
		"Delivery Signature": False,
		"Additional Handling Surcharge": False,
		#This is false since it's calculated after everything
		"Fuel Surcharge": False,
	}

def add_fuel_surcharge(charge_type):
    return add_fuel_surcharge_index[charge_type]

def calc_fuel_surcharge(date, total_billed_amount, service_level):
    # date is a string 'mm/dd/yyyy'
    # add_fuel_surcharge_index has True for those charge types where the total
    # is calculated with for the percentage calculation with fuel shortage
    fedex_total = 0
    fuel_dic = {"Charge Type": "Fuel Surcharge", }
    date_dic = convert_date_to_monday(date)

    year = date_dic["year"]
    month = date_dic["month"]
    day = date_dic["day"]

    delivery_type = fedex_delivery_type[service_level]

    fuel_surcharge_percent = get_fuel_rate(year, month, day, delivery_type) / 100.0

    return fuel_surcharge_percent * total_billed_amount



def get_fuel_rate(year, month, day, delivery_type):
    date_string = str(month) + "/" + str(day) + "/" + str(year)
    fuel_list = fuel_rate_index[date_string]
    if delivery_type == "Express":
        return float(fuel_list[0])
    elif delivery_type == "Ground":
        return float(fuel_list[1])
    else:
        msg = "ERROR: delivery type " + delivery_type + " is seen."
        raise Exception(msg)


def convert_date_to_monday(date_string):
    # Format will be in mm/dd/yyyy format
    date_re_pattern = "(?P<month>\d+)\\(?P<day>\d+)\\(?P<year>\d+)"
    r = re.match(r"(?P<month>\d+)\/(?P<day>\d+)\/(?P<year>\d+)", date_string)
    month = int(r.group('month'))
    year = int(r.group('year'))
    day = int(r.group('day'))

    original_date = datetime.date(year, month, day)
    weekday = original_date.weekday()
    date_diff = datetime.timedelta(days=weekday)
    new_date = original_date - date_diff

    return {"year": new_date.year, "month": new_date.month, "day": new_date.day, }

# Fuel rate percentages are stored in [Express_value, Ground_value]

fuel_rate_index = {
    # Each date is for the entire week.

    "1/2/2017": [2.50, 4.00],
    "1/9/2017": [2.50, 4.00],
    "1/16/2017": [2.50, 4.00],
    "1/23/2017": [2.50, 4.00],
    "1/30/2017": [2.50, 4.00],
    "2/6/2017": [3.50, 4.50],
    "2/13/2017": [3.50, 4.25],
    "2/20/2017": [3.50, 4.50],
    "2/27/2017": [3.75, 4.50],
    "3/6/2017": [3.75, 4.50],
    "3/13/2017": [3.50, 4.50],
    "3/20/2017": [3.25, 4.50],
    "3/27/2017": [2.75, 4.25],
    "4/3/2017": [2.75, 4.25],
    "4/10/2017": [3.00, 4.25],
    "4/17/2017": [3.50, 4.50],
    "4/24/2017": [3.75, 4.50],
    "5/1/2017": [3.50, 4.50],
    "5/8/2017": [3.00, 4.50],
    "5/15/2017": [2.50, 4.50],
}

fedex_delivery_type = {
    'Priority Overnight': "Express",
    'Standard Overnight': "Express",
    '2 Day AM': "Express",
    '2 Day': "Express",
    'Express Saver (3 Day)': "Express",
    'Ground': "Ground",
    'Home Delivery': "Ground",
    'Smart Post 1-16 oz': "Ground",
    'Smart Post 1-70 lbs': "Ground",
	}
