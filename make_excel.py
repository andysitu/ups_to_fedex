import openpyxl
import ffile

alphabet_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'K', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

def output_raw_data(raw_ups_data):
	# outputs raw ups data into excel file
	ffile.move_dir("excel")

	wb = openpyxl.Workbook()
	sheet = wb.active
	i = 1

	for tracking_num, data_types in raw_ups_data.items():
		sheet['A' + str(i)] = tracking_num
		i += 1
		for data_type, data_list in data_types.items():
			sheet['A' + str(i)] = data_type
			i += 1
			for data in data_list:
				sheet['A' + str(i)] = str(data)
				# for j, key in enumerate(data):
				# 	sheet[alphabet_list[j * 2] + str(i)] = key
				# 	sheet[alphabet_list[j * 2 + 1] + str(i)] = str(data[key])
				i += 1
	wb.save('ups_raw_data.xlsx')

	ffile.dir_back()

def output_conv_ups_data(ups_data):
	# outputs ups data into excel file
	ffile.move_dir("excel")
	
	wb = openpyxl.Workbook()
	sheet = wb.active
	i = 1

	for date, data_dic in ups_data.items():
		sheet['A' + str(i)] = date
		i += 1
		for tracking_num, data_obj in data_dic.items():
			sheet['A' + str(i)] = tracking_num
			i += 1

			simple_data_list = data_obj.get_simple_datalist_str()
			for s in simple_data_list:
				sheet['A' + str(i)] = s
				i += 1

			sheet['A' + str(i)] = "detail"
			i += 1
			detail_data_list = data_obj.get_detail_datalist_str()
			for d in detail_data_list:
				sheet['A' + str(i)] = d
				i += 1
	wb.save('ups_data.xlsx')

	ffile.dir_back()

def output_fedex_ups_dat(fedex_list_inst):
	dates_list =  fedex_list_inst.get_dates()

	ffile.move_dir("excel")
	
	wb = openpyxl.Workbook()
	sheet = wb.active
	i = 1

	sheet.freeze_panes = 'A2'

	UPS_column_width = 30
	tracking_num_width = 20
	fedex_column_width = 30
	sheet.column_dimensions['D'].width = UPS_column_width
	sheet.column_dimensions['E'].width = tracking_num_width
	sheet.column_dimensions['F'].width = fedex_column_width


	header_list = ["Date", "Zone", "Weight", "UPS", "Track Num/Rate", "Fedex", "FX Rate"]

	date_col_in = header_list.index("Date")
	date_col_letter = alphabet_list[date_col_in]

	zone_col_in = header_list.index("Zone")
	zone_col_letter = alphabet_list[zone_col_in]

	weight_col_in = header_list.index("Weight")
	weight_col_letter = alphabet_list[weight_col_in]

	ups_col_in = header_list.index("UPS")
	ups_col_letter = alphabet_list[ups_col_in]

	ups_rate_col_in = header_list.index("Track Num/Rate")
	ups_rate_col_letter = alphabet_list[ups_rate_col_in]

	fedex_col_in = header_list.index("Fedex")
	fedex_col_letter = alphabet_list[fedex_col_in]

	fedex_rate_col_in = header_list.index("FX Rate")
	fedex_rate_col_letter = alphabet_list[fedex_rate_col_in]

	for column_num, header in enumerate(header_list):
		column_letter = alphabet_list[column_num]
		sheet[column_letter + str(i)] = header
	i += 1

	for date in dates_list:
		tracking_num_list = fedex_list_inst.get_ups_tracking_nums(date)
		for tracking_num in tracking_num_list:
			data_dic = fedex_list_inst.get_first_dataset(date, tracking_num)
			zone = data_dic["Zone"]
			date = data_dic["Date"]
			weight = data_dic["Weight"]

			sheet[date_col_letter + str(i)] = date
			sheet[zone_col_letter + str(i)] = zone
			sheet[weight_col_letter + str(i)] = weight
			sheet[ups_col_letter + str(i)] = "UPS"
			sheet[ups_rate_col_letter + str(i)] = tracking_num
			sheet[fedex_col_letter + str(i)] = "Fedex"
			sheet[date_col_letter + str(i)] = ""

			i += 1

			ups_data_dic = data_dic["ups"]
			fedex_data_dic = data_dic["fedex"]

			total_ups_rate = ups_data_dic["Total Charge"]
			total_fedex_rate = fedex_data_dic["Total Charge"]

			ups_data_list = ups_data_dic['Charges List']
			fedex_data_list = fedex_data_dic['Charges List']

			for list_index, ups_rate_dic in enumerate(ups_data_list):
				# print(ups_data_list, len(ups_data_list))
				# print(fedex_data_list, len(fedex_data_list))
				fedex_rate_dic = fedex_data_list[list_index]

				ups_rate = ups_rate_dic["Billed Charge"]
				ups_charge_type = ups_rate_dic["Charge Type"]

				fedex_rate = fedex_rate_dic["Billed Charge"]
				fedex_charge_type = fedex_rate_dic["Charge Type"]

				sheet[ups_col_letter + str(i)] = ups_charge_type
				sheet[ups_rate_col_letter + str(i)] = ups_rate
				sheet[fedex_col_letter + str(i)] = fedex_charge_type
				sheet[fedex_rate_col_letter + str(i)] = fedex_rate
				i += 1

			sheet[ups_col_letter + str(i)] = "TOTAL"
			sheet[ups_rate_col_letter + str(i)] = total_ups_rate
			sheet[fedex_col_letter + str(i)] = "TOTAL"
			sheet[fedex_rate_col_letter + str(i)] = total_fedex_rate

			diff_word_in = fedex_rate_col_in + 1
			diff_word_col_letter = alphabet_list[diff_word_in]
			diff_rate_in = fedex_rate_col_in + 2
			diff_rate_col_letter = alphabet_list[diff_rate_in]
			
			sheet[diff_word_col_letter + str(i)] = "Difference"
			sheet[diff_rate_col_letter + str(i)] = total_ups_rate - total_fedex_rate

			i += 1

			# print(ups_data_list)
			# for ups_data_index in ups_data_list:
			# 	ups_charges_list = ups_data_list[ups_data_index]["Charges List"]
			# 	fedex_charges_list =fedex_data_list[ups_data_index]["Chares List"]
			# 	#UPS & Fedex should have the same keys
			# 	print(ups_charges_list)
			# 	print(fedex_charges_list)

	wb.save('ups_and_fedex_data.xlsx')

	ffile.dir_back()